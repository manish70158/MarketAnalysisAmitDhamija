"""
NSE Participant-wise Open Interest Dashboard
Uses ACTUAL NSE data - exact replica of the screenshot format
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import requests
import time
import os
from io import StringIO


class NSEParticipantOIDashboard:
    def __init__(self):
        # Colors matching screenshot
        self.green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        self.red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        self.yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        self.dark_header = PatternFill(start_color="002060", end_color="002060", fill_type="solid")

        self.white_font = Font(color="FFFFFF", bold=True, size=10)
        self.bold_font = Font(bold=True, size=10)
        self.normal_font = Font(size=10)
        self.small_font = Font(size=8)

        thin_side = Side(style='thin', color='000000')
        self.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        self.center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left = Alignment(horizontal='left', vertical='center')
        self.right = Alignment(horizontal='right', vertical='center')

    # ──────────────────── NSE Data Fetching ────────────────────

    def _get_session(self):
        """Create a requests session with NSE cookies."""
        if not hasattr(self, '_session') or self._session is None:
            self._session = requests.Session()
            self._session.headers.update({
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
                'Accept-Language': 'en-US,en;q=0.9',
                'Referer': 'https://www.nseindia.com/',
            })
            # Hit NSE homepage to get cookies
            try:
                self._session.get("https://www.nseindia.com/", timeout=10)
            except Exception:
                pass
        return self._session

    def fetch_nse_data(self, date_str):
        """
        Fetch participant OI data from NSE archives.
        Always fetches fresh from NSE; uses local cache only as fallback.
        date_str format: 'DDMMYYYY' (e.g., '24072026')
        Returns: dict with participant data or None
        """
        csv_file = f"participant_oi_{date_str}.csv"
        url = f"https://archives.nseindia.com/content/nsccl/fao_participant_oi_{date_str}.csv"

        # Try downloading fresh data from NSE
        try:
            session = self._get_session()
            r = session.get(url, timeout=15)
            if r.status_code == 200 and len(r.content) > 100:
                with open(csv_file, 'wb') as f:
                    f.write(r.content)
                print(f"  Downloaded: {csv_file}")
                return self.parse_nse_csv(r.text)
            else:
                print(f"  Not available: {date_str} (status {r.status_code})")
        except Exception as e:
            print(f"  Error fetching {date_str}: {e}")

        # Fallback: use local cache if download failed
        if os.path.exists(csv_file):
            print(f"  Using cached: {csv_file}")
            with open(csv_file, 'r') as f:
                return self.parse_nse_csv(f.read())

        return None

    def parse_nse_csv(self, csv_text):
        """
        Parse NSE participant OI CSV into structured dict
        Returns: {
            'Client': {'fut_idx_long': X, 'fut_idx_short': X, ...},
            'DII': {...}, 'FII': {...}, 'Pro': {...}
        }
        """
        lines = csv_text.strip().split('\n')
        # Skip header line (first line is title, second is column headers)
        data = {}

        for line in lines[2:]:  # Skip title and header
            parts = [p.strip().strip('"') for p in line.split(',')]
            if len(parts) < 14:
                continue

            participant = parts[0].strip()
            if participant in ('Client', 'DII', 'FII', 'Pro', 'TOTAL'):
                data[participant] = {
                    'fut_idx_long': int(parts[1]),
                    'fut_idx_short': int(parts[2]),
                    'fut_stk_long': int(parts[3]),
                    'fut_stk_short': int(parts[4].strip()),
                    'opt_idx_call_long': int(parts[5]),
                    'opt_idx_put_long': int(parts[6]),
                    'opt_idx_call_short': int(parts[7]),
                    'opt_idx_put_short': int(parts[8]),
                    'opt_stk_call_long': int(parts[9]),
                    'opt_stk_put_long': int(parts[10]),
                    'opt_stk_call_short': int(parts[11]),
                    'opt_stk_put_short': int(parts[12]),
                    'total_long': int(parts[13].strip()),
                    'total_short': int(parts[14]),
                }

        return data if data else None

    def get_trading_dates(self):
        """Get last 3 trading dates (try multiple dates to handle weekends/holidays)"""
        # Try dates in reverse order to find 3 valid trading days
        from datetime import timedelta
        today = datetime.now()

        dates_to_try = []
        for i in range(10):  # Try last 10 days
            d = today - timedelta(days=i)
            dates_to_try.append(d.strftime('%d%m%Y'))

        valid_dates = []
        for date_str in dates_to_try:
            data = self.fetch_nse_data(date_str)
            if data:
                valid_dates.append((date_str, data))
                if len(valid_dates) == 3:
                    break
            time.sleep(0.5)

        return valid_dates

    # ──────────────────── Number Formatting ────────────────────

    def fmt(self, num):
        """Format number in Indian system (e.g., 1,22,496)"""
        if num == 0:
            return "0"
        if num < 0:
            return "-" + self.fmt(abs(num))
        s = str(int(num))
        if len(s) <= 3:
            return s
        last3 = s[-3:]
        rest = s[:-3]
        parts = []
        while len(rest) > 2:
            parts.insert(0, rest[-2:])
            rest = rest[:-2]
        if rest:
            parts.insert(0, rest)
        return ",".join(parts) + "," + last3

    # ──────────────────── Excel Cell Helpers ────────────────────

    def cell(self, ws, row, col, value, fill=None, font=None, align=None):
        c = ws.cell(row=row, column=col)
        c.value = value
        c.border = self.border
        if fill:
            c.fill = fill
        c.font = font or self.normal_font
        c.alignment = align or self.left
        return c

    def merged_row(self, ws, row, start_col, end_col, value, fill=None, font=None):
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
        c = ws.cell(row=row, column=start_col)
        c.value = value
        c.alignment = self.center
        c.font = font or self.bold_font
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).border = self.border
            if fill:
                ws.cell(row=row, column=col).fill = fill

    # ──────────────────── Section Builders ────────────────────

    def build_title(self, ws, row):
        self.merged_row(ws, row, 1, 10,
                        "Telegram Channel : t.me/MarketAnalysisNiftyBankNifty",
                        self.dark_header, self.white_font)
        # Fill right side with same style
        for col in range(11, 15):
            self.cell(ws, row, col, "", self.dark_header, self.white_font, self.center)
        return row + 1

    def build_subtitle(self, ws, row):
        self.merged_row(ws, row, 1, 10,
                        "Oil Changes  |  Youtube Channel  |  Market Analysis With Amit Dhamija",
                        self.yellow_fill, self.bold_font)
        for col in range(11, 15):
            self.cell(ws, row, col, "", self.yellow_fill, self.bold_font, self.center)
        return row + 1

    def build_main_header(self, ws, row):
        self.merged_row(ws, row, 1, 7, "Futures & Options", None, self.bold_font)
        self.merged_row(ws, row, 8, 10, "Total Positions Carried", None, self.bold_font)
        self.cell(ws, row, 11, "", None, None, self.center)
        # Right side header - merge into one row
        ws.merge_cells(start_row=row, start_column=12, end_row=row, end_column=14)
        c = ws.cell(row=row, column=12)
        c.value = "Positions Bought / Sold Today"
        c.font = self.bold_font
        c.alignment = self.center
        c.border = self.border
        ws.cell(row=row, column=13).border = self.border
        ws.cell(row=row, column=14).border = self.border
        return row + 1

    def build_section_header(self, ws, row, longs_title, shorts_title):
        self.cell(ws, row, 1, "", self.yellow_fill, self.bold_font, self.center)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        self.cell(ws, row, 2, longs_title, self.yellow_fill, self.bold_font, self.center)
        ws.cell(row=row, column=3).border = self.border
        ws.cell(row=row, column=3).fill = self.yellow_fill
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
        self.cell(ws, row, 4, shorts_title, self.yellow_fill, self.bold_font, self.center)
        ws.cell(row=row, column=5).border = self.border
        ws.cell(row=row, column=5).fill = self.yellow_fill
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
        self.cell(ws, row, 6, "Net Buy / Sell for Today", self.yellow_fill, self.bold_font, self.center)
        ws.cell(row=row, column=7).border = self.border
        ws.cell(row=row, column=7).fill = self.yellow_fill
        self.cell(ws, row, 8, "TODAY", self.yellow_fill, self.bold_font, self.center)
        self.cell(ws, row, 9, "1 DAY AGO", self.yellow_fill, self.bold_font, self.center)
        self.cell(ws, row, 10, "2 DAYS AGO", self.yellow_fill, self.bold_font, self.center)
        for col in range(11, 15):
            self.cell(ws, row, col, "", self.yellow_fill, self.bold_font, self.center)
        return row + 1

    def build_data_row(self, ws, row, participant, long_change, short_change,
                       today_oi, day1_oi, day2_oi, right_data=None, is_put=False):
        """Build one participant data row — uses RED/GREEN TEXT (no background fill)"""
        # Font colors
        green_text = Font(color="008000", bold=True, size=10)
        red_text = Font(color="FF0000", bold=True, size=10)

        # Actions
        long_action = "Added Longs" if long_change >= 0 else "Closed Longs"
        short_action = "Added Shorts" if short_change >= 0 else "Closed Shorts"

        if not is_put:
            # CALL/FUTURES: Added Longs=GREEN text, Closed Longs=RED text
            long_font = green_text if long_change >= 0 else red_text
            # Added Shorts=RED text, Closed Shorts=GREEN text
            short_font = red_text if short_change >= 0 else green_text
        else:
            # PUT: Added Longs=RED text (buying puts=bearish)
            long_font = red_text if long_change >= 0 else green_text
            # Added Shorts=GREEN text (selling puts=bullish)
            short_font = green_text if short_change >= 0 else red_text

        # Net = Long change - Short change
        net_value = long_change - short_change
        if not is_put:
            net_bullish = net_value > 0
        else:
            net_bullish = net_value < 0
        net_font = green_text if net_bullish else red_text
        net_label = "Bought Net" if net_value > 0 else "Sold Net"

        # A: Participant
        self.cell(ws, row, 1, participant, None, self.bold_font, self.left)
        # B: Longs Action (text color)
        self.cell(ws, row, 2, long_action, None, long_font, self.center)
        # C: Longs Value (text color)
        self.cell(ws, row, 3, self.fmt(long_change), None, long_font, self.right)
        # D: Shorts Action (text color)
        self.cell(ws, row, 4, short_action, None, short_font, self.center)
        # E: Shorts Value (text color)
        self.cell(ws, row, 5, self.fmt(short_change), None, short_font, self.right)
        # F: Net Label (text color)
        self.cell(ws, row, 6, net_label, None, net_font, self.center)
        # G: Net Value (text color)
        self.cell(ws, row, 7, self.fmt(net_value), None, net_font, self.right)
        # H: TODAY total OI
        self.cell(ws, row, 8, self.fmt(today_oi), None, self.normal_font, self.right)
        # I: 1 DAY AGO
        self.cell(ws, row, 9, self.fmt(day1_oi), None, self.normal_font, self.right)
        # J: 2 DAYS AGO
        self.cell(ws, row, 10, self.fmt(day2_oi), None, self.normal_font, self.right)
        # K: gap
        self.cell(ws, row, 11, "", None, None, self.center)

        # Right side
        if right_data and right_data.get('category'):
            rv = right_data['value']
            r_fill = self.green_fill if rv > 0 else self.red_fill
            r_label = "Bought Net" if rv > 0 else "Sold Net"
            self.cell(ws, row, 12, r_label, r_fill, self.bold_font, self.center)
            self.cell(ws, row, 13, right_data['category'], None, self.bold_font, self.center)
            self.cell(ws, row, 14, self.fmt(rv), r_fill, self.bold_font, self.right)
        else:
            for col in range(12, 15):
                self.cell(ws, row, col, "", None, None, self.center)

        return row + 1

    def build_total_row(self, ws, row, today_total, day1_total, day2_total):
        self.cell(ws, row, 1, "Total", self.yellow_fill, self.bold_font, self.center)
        for col in range(2, 8):
            self.cell(ws, row, col, "", self.yellow_fill, self.bold_font, self.center)
        self.cell(ws, row, 8, self.fmt(today_total), self.yellow_fill, self.bold_font, self.right)
        self.cell(ws, row, 9, self.fmt(day1_total), self.yellow_fill, self.bold_font, self.right)
        self.cell(ws, row, 10, self.fmt(day2_total), self.yellow_fill, self.bold_font, self.right)
        for col in range(11, 15):
            self.cell(ws, row, col, "", self.yellow_fill, self.bold_font, self.center)
        return row + 1

    def build_separator(self, ws, row):
        self.merged_row(ws, row, 1, 10,
                        "YouTube Channel  |  Market Analysis  |  With Amit Dhamija",
                        self.yellow_fill, self.bold_font)
        # Keep right-side columns independent (not merged)
        for col in range(11, 15):
            self.cell(ws, row, col, "", self.yellow_fill, self.bold_font, self.center)
        return row + 1

    # ──────────────────── Section Creator ────────────────────

    def build_section(self, ws, row, longs_title, shorts_title,
                      long_key, short_key, today_data, day1_data, day2_data,
                      right_data_list, is_put=False):
        """Build a complete section using actual NSE data"""
        row = self.build_section_header(ws, row, longs_title, shorts_title)

        participants = ['Client', 'DII', 'FII', 'Pro']

        # Calculate TOTAL OI for the Total row (sum of all Long OI)
        today_total_oi = sum(today_data[p][long_key] for p in participants)
        day1_total_oi = sum(day1_data[p][long_key] for p in participants)
        day2_total_oi = sum(day2_data[p][long_key] for p in participants)

        for i, p in enumerate(participants):
            # Get OI values
            t_long = today_data[p][long_key]
            t_short = today_data[p][short_key]
            d1_long = day1_data[p][long_key]
            d1_short = day1_data[p][short_key]
            d2_long = day2_data[p][long_key]
            d2_short = day2_data[p][short_key]

            # Calculate changes (today vs 1 day ago)
            long_change = t_long - d1_long
            short_change = t_short - d1_short

            # Total positions carried = Net Position (Long - Short) for that date
            today_oi = t_long - t_short
            day1_oi = d1_long - d1_short
            day2_oi = d2_long - d2_short

            # Right side data
            r_data = right_data_list[i] if i < len(right_data_list) else None

            # Display participant name as in screenshot
            display_name = "Clients" if p == "Client" else p

            row = self.build_data_row(ws, row, display_name,
                                      long_change, short_change,
                                      today_oi, day1_oi, day2_oi, r_data,
                                      is_put=is_put)

        # Total row shows aggregate Long OI across all participants
        row = self.build_total_row(ws, row, today_total_oi, day1_total_oi, day2_total_oi)
        row = self.build_separator(ws, row)
        return row

    # ──────────────────── Main Dashboard Creator ────────────────────

    def create_dashboard(self, output_filename):
        print("Fetching NSE Participant-wise OI data...")
        dates = self.get_trading_dates()

        if len(dates) < 3:
            print(f"ERROR: Need at least 3 trading days of data, got {len(dates)}")
            print("Please ensure CSV files are available for recent dates.")
            return False

        today_date, today_data = dates[0]
        day1_date, day1_data = dates[1]
        day2_date, day2_data = dates[2]

        print(f"\n  TODAY:      {today_date[:2]}-{today_date[2:4]}-{today_date[4:]}")
        print(f"  1 DAY AGO:  {day1_date[:2]}-{day1_date[2:4]}-{day1_date[4:]}")
        print(f"  2 DAYS AGO: {day2_date[:2]}-{day2_date[2:4]}-{day2_date[4:]}")

        # Compute right-side: Per-participant TOTAL net bought/sold across ALL instruments
        right_participant_net = {}
        for p in ['Client', 'DII', 'FII', 'Pro']:
            total_long_chg = today_data[p]['total_long'] - day1_data[p]['total_long']
            total_short_chg = today_data[p]['total_short'] - day1_data[p]['total_short']
            right_participant_net[p] = total_long_chg - total_short_chg

        # Compute per-instrument totals for right side labels
        right_instrument = {}
        for key_long, key_short, label in [
            ('fut_idx_long', 'fut_idx_short', 'Index Fut'),
            ('opt_idx_call_long', 'opt_idx_call_short', 'Index Calls'),
            ('opt_idx_put_long', 'opt_idx_put_short', 'Index Puts'),
            ('fut_stk_long', 'fut_stk_short', 'Stock Fut'),
            ('opt_stk_call_long', 'opt_stk_call_short', 'Stock Calls'),
            ('opt_stk_put_long', 'opt_stk_put_short', 'Stock Puts'),
        ]:
            for p in ['Client', 'DII', 'FII', 'Pro']:
                lc = today_data[p][key_long] - day1_data[p][key_long]
                sc = today_data[p][key_short] - day1_data[p][key_short]
                right_instrument[(p, label)] = lc - sc

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Participant OI"

        row = 1
        row = self.build_title(ws, row)
        row = self.build_subtitle(ws, row)
        row = self.build_main_header(ws, row)

        # Right-side: Per-participant net for each instrument type
        instrument_labels = ['Index Fut', 'Index Calls', 'Index Puts',
                             'Stock Fut', 'Stock Calls', 'Stock Puts']
        participants = ['Client', 'DII', 'FII', 'Pro']

        def make_right_list(inst_label):
            """Create right-side data showing per-participant net for an instrument"""
            return [
                {'category': inst_label, 'value': right_instrument[(p, inst_label)]}
                for p in participants
            ]

        # Also compute each participant's TOTAL net across all instruments
        # (for the right-side summary at the bottom)
        participant_totals = {}
        for p in participants:
            total_net = 0
            for inst in instrument_labels:
                total_net += right_instrument[(p, inst)]
            participant_totals[p] = total_net

        # Build all 6 sections (no right-side data per row — will fill right side separately)
        empty_right = [None, None, None, None]

        row = self.build_section(ws, row, "Index Future Longs", "Index Future Shorts",
                                 'fut_idx_long', 'fut_idx_short',
                                 today_data, day1_data, day2_data, empty_right)

        row = self.build_section(ws, row, "Index Call Longs", "Index Call Shorts",
                                 'opt_idx_call_long', 'opt_idx_call_short',
                                 today_data, day1_data, day2_data, empty_right)

        # PUT sections have INVERTED colors (buying puts = bearish)
        row = self.build_section(ws, row, "Index Put Longs", "Index Put Shorts",
                                 'opt_idx_put_long', 'opt_idx_put_short',
                                 today_data, day1_data, day2_data, empty_right, is_put=True)

        row = self.build_section(ws, row, "Stock Future Longs", "Stock Future Shorts",
                                 'fut_stk_long', 'fut_stk_short',
                                 today_data, day1_data, day2_data, empty_right)

        row = self.build_section(ws, row, "Stock Calls Longs", "Stock Calls Shorts",
                                 'opt_stk_call_long', 'opt_stk_call_short',
                                 today_data, day1_data, day2_data, empty_right)

        row = self.build_section(ws, row, "Stock Puts Longs", "Stock Puts Shorts",
                                 'opt_stk_put_long', 'opt_stk_put_short',
                                 today_data, day1_data, day2_data, empty_right, is_put=True)

        # === RIGHT SIDE: "Positions Bought / Sold Today" ===
        # 4 vertical sub-sections: Clients, DII, FII, Pro
        # Uses RED/GREEN TEXT (font color) — NOT background fill (matching screenshot)
        green_text = Font(color="008000", bold=True, size=10)
        red_text = Font(color="FF0000", bold=True, size=10)

        right_row = 4  # Start at row 4

        for p in participants:
            display_name = "Clients" if p == "Client" else p

            # Participant header row
            self.cell(ws, right_row, 12, display_name, None, self.bold_font, self.center)
            self.cell(ws, right_row, 13, "", None, self.normal_font, self.center)
            self.cell(ws, right_row, 14, "", None, self.normal_font, self.center)
            right_row += 1

            # One row per instrument showing net bought/sold (text color only)
            for inst in instrument_labels:
                val = right_instrument[(p, inst)]
                font_color = green_text if val > 0 else red_text
                label = "Bought Net" if val > 0 else "Sold Net"
                self.cell(ws, right_row, 12, label, None, font_color, self.center)
                self.cell(ws, right_row, 13, inst, None, self.normal_font, self.center)
                self.cell(ws, right_row, 14, self.fmt(val), None, font_color, self.right)
                right_row += 1

            # Total row for this participant
            total_val = participant_totals[p]
            t_font = green_text if total_val > 0 else red_text
            t_label = "Bought Net" if total_val > 0 else "Sold Net"
            self.cell(ws, right_row, 12, t_label, None, t_font, self.center)
            self.cell(ws, right_row, 13, "TOTAL", None, self.bold_font, self.center)
            self.cell(ws, right_row, 14, self.fmt(total_val), None, t_font, self.right)
            right_row += 1

            # Blank spacing row
            right_row += 1

        # Bottom info
        row += 1
        info_lines = [
            "Notes : In System & Result",
            "1) World Market Should Have Opposite (Trend)",
            "Watch out for market impacting news",
            "POWER PACKED WEBINAR PACKAGE for Best Strategies:",
            "Telegram : Market Analysis With Amit Dhamija OR",
            "t.me/MarketAnalysisNiftyBankNifty",
        ]
        for text in info_lines:
            self.cell(ws, row, 1, text, None, self.small_font, self.left)
            row += 1

        # Column widths
        for col, w in enumerate([9, 13, 11, 14, 11, 11, 11, 11, 11, 11, 2, 11, 11, 12], start=1):
            ws.column_dimensions[get_column_letter(col)].width = w

        ws.row_dimensions[1].height = 20

        wb.save(output_filename)
        print(f"\nDashboard created: {output_filename}")
        print(f"  Data: ACTUAL NSE Participant-wise OI")
        print(f"  Format: Exact screenshot match")
        return True


def main():
    dashboard = NSEParticipantOIDashboard()
    date_str = datetime.now().strftime("%d_%b_%Y").upper()
    output_filename = f"nse_participant_oi_{date_str}.xlsx"
    dashboard.create_dashboard(output_filename)


if __name__ == "__main__":
    main()
