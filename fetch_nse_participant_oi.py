"""
NSE F&O Participant-wise Open Interest Data Fetcher
Downloads and formats participant OI data from NSE India
"""

import requests
import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import time
import os
import sys
from pathlib import Path

class NSEParticipantOIFetcher:
    def __init__(self):
        self.base_url = "https://www.nseindia.com"
        self.session = requests.Session()
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.9',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Referer': 'https://www.nseindia.com/',
        }

    def get_cookies(self):
        """Visit NSE homepage to get cookies"""
        try:
            response = self.session.get(
                self.base_url,
                headers=self.headers,
                timeout=10
            )
            return response.status_code == 200
        except Exception as e:
            print(f"Error getting cookies: {e}")
            return False

    def download_participant_oi(self, date_str):
        """
        Download participant-wise OI data for a specific date
        date_str format: 'DD-MMM-YYYY' (e.g., '24-JUL-2026')
        """
        # First get cookies
        if not self.get_cookies():
            print("Failed to get cookies from NSE")
            return None

        time.sleep(2)  # Wait before making actual request

        # Format date for URL (DDMMMYYYY)
        date_formatted = date_str.replace('-', '').upper()

        # URL pattern for participant OI CSV
        csv_url = f"https://archives.nseindia.com/content/nsccl/fao_participant_oi_{date_formatted}.csv"

        try:
            response = self.session.get(
                csv_url,
                headers=self.headers,
                timeout=30
            )

            if response.status_code == 200:
                print(f"✓ Downloaded data for {date_str}")
                return response.content
            else:
                print(f"✗ Failed to download. Status code: {response.status_code}")
                return None

        except Exception as e:
            print(f"Error downloading data: {e}")
            return None

    def parse_csv_data(self, csv_content):
        """Parse the NSE CSV file"""
        try:
            from io import StringIO
            csv_string = csv_content.decode('utf-8')

            # Read CSV - NSE files often have metadata rows at top
            df = pd.read_csv(StringIO(csv_string))

            return df
        except Exception as e:
            print(f"Error parsing CSV: {e}")
            return None

    def determine_action_and_view(self, row):
        """
        Determine the action type and market view based on OI changes
        Returns: (action, is_bullish)
        """
        # Get the OI change values
        long_oi = float(row.get('Long OI', 0) or 0)
        short_oi = float(row.get('Short OI', 0) or 0)

        # Determine action based on OI changes
        if long_oi > 0 and short_oi > 0:
            if long_oi > short_oi:
                return "Build Long", True
            else:
                return "Build Short", False
        elif long_oi > 0 and short_oi < 0:
            return "Closed Shorts", True
        elif long_oi < 0 and short_oi > 0:
            return "Closed Longs", False
        elif long_oi < 0 and short_oi < 0:
            if abs(long_oi) > abs(short_oi):
                return "Closed Longs", False
            else:
                return "Closed Shorts", True
        else:
            return "Build Net", long_oi >= 0

    def format_excel(self, df, output_file):
        """Format the data into Excel with colors"""
        try:
            # Create a new workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Participant OI"

            # Define colors
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            green_font = Font(color="006100", bold=True)
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            red_font = Font(color="9C0006", bold=True)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            # Add headers
            headers = list(df.columns)
            ws.append(headers)

            # Format header row
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Add data rows
            for idx, row in df.iterrows():
                ws.append(list(row))

                # Determine if row is bullish or bearish
                action, is_bullish = self.determine_action_and_view(row)

                # Format the entire row based on view
                current_row = ws.max_row
                for cell in ws[current_row]:
                    if is_bullish:
                        cell.font = green_font
                        cell.fill = green_fill
                    else:
                        cell.font = red_font
                        cell.fill = red_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # Adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save workbook
            wb.save(output_file)
            print(f"✓ Excel file created: {output_file}")
            return True

        except Exception as e:
            print(f"Error creating Excel: {e}")
            return False


def process_local_csv(csv_file_path, date_str="24-JUL-2026"):
    """Process a locally downloaded CSV file"""
    fetcher = NSEParticipantOIFetcher()

    print(f"\n{'='*60}")
    print(f"Processing local CSV file")
    print(f"{'='*60}\n")
    print(f"File: {csv_file_path}")

    try:
        with open(csv_file_path, 'rb') as f:
            csv_content = f.read()

        # Parse the CSV
        df = fetcher.parse_csv_data(csv_content)

        if df is not None:
            print(f"✓ Data parsed successfully. Rows: {len(df)}, Columns: {len(df.columns)}")

            # Display first few rows
            print("\nSample data:")
            print(df.head())
            print(f"\nColumns: {list(df.columns)}")

            # Create formatted Excel file
            output_excel = f"participant_oi_{date_str.replace('-', '_')}_formatted.xlsx"
            fetcher.format_excel(df, output_excel)

            print(f"\n{'='*60}")
            print(f"✓ Process completed successfully!")
            print(f"✓ Output: {output_excel}")
            print(f"{'='*60}\n")
            return True
        else:
            print("✗ Failed to parse CSV data")
            return False

    except FileNotFoundError:
        print(f"✗ File not found: {csv_file_path}")
        return False
    except Exception as e:
        print(f"✗ Error processing file: {e}")
        return False


def main():
    """Main function to fetch and process NSE participant OI data"""

    # Initialize fetcher
    fetcher = NSEParticipantOIFetcher()

    # Start date: 24-July-2026
    start_date = "24-JUL-2026"

    print(f"\n{'='*60}")
    print(f"NSE F&O Participant-wise Open Interest Fetcher")
    print(f"{'='*60}\n")

    # Check if user provided a local file
    if len(sys.argv) > 1:
        csv_file = sys.argv[1]
        if os.path.exists(csv_file):
            print(f"Processing local file: {csv_file}\n")
            process_local_csv(csv_file, start_date)
            return
        else:
            print(f"File not found: {csv_file}")
            return

    print(f"Fetching data for: {start_date}")
    print(f"Source: NSE India Archives\n")

    # Download the CSV file
    csv_content = fetcher.download_participant_oi(start_date)

    if csv_content:
        # Save raw CSV for reference
        raw_csv_file = f"participant_oi_{start_date.replace('-', '_')}_raw.csv"
        with open(raw_csv_file, 'wb') as f:
            f.write(csv_content)
        print(f"✓ Raw CSV saved: {raw_csv_file}")

        # Parse the CSV
        df = fetcher.parse_csv_data(csv_content)

        if df is not None:
            print(f"✓ Data parsed successfully. Rows: {len(df)}")

            # Display first few rows
            print("\nSample data:")
            print(df.head())

            # Create formatted Excel file
            output_excel = f"participant_oi_{start_date.replace('-', '_')}_formatted.xlsx"
            fetcher.format_excel(df, output_excel)

            print(f"\n{'='*60}")
            print(f"✓ Process completed successfully!")
            print(f"{'='*60}\n")
        else:
            print("✗ Failed to parse CSV data")
    else:
        print("✗ Failed to download data automatically")
        print("\n" + "="*60)
        print("Manual Download Instructions:")
        print("="*60)
        print("\n1. Visit: https://www.nseindia.com/all-reports-derivatives")
        print("2. Look for 'F&O-Participant wise Open Interest (csv)'")
        print("3. Download the file for 24-JUL-2026")
        print("4. Run this script with the file path:")
        print(f"   python {sys.argv[0]} <path-to-downloaded-csv>")
        print("\n" + "="*60)


if __name__ == "__main__":
    main()
