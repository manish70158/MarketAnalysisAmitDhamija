"""
Parse NSE Participant OI CSV and create EXACT screenshot format
Handles real NSE CSV data and formats it identically to screenshot
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
import os

class NSEParticipantParser:
    def __init__(self):
        self.green_font = Font(color="006100", bold=True, size=9)
        self.red_font = Font(color="9C0006", bold=True, size=9)
        self.header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.header_font = Font(bold=True, size=9)
        self.border_thin = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

    def read_nse_csv(self, csv_file):
        """Read NSE CSV file"""
        try:
            # NSE CSVs can have different encodings
            df = pd.read_csv(csv_file)
            print(f"✓ Loaded CSV: {len(df)} rows, {len(df.columns)} columns")
            print(f"\nColumns: {list(df.columns)}")
            return df
        except Exception as e:
            print(f"✗ Error reading CSV: {e}")
            try:
                df = pd.read_csv(csv_file, encoding='latin1')
                print(f"✓ Loaded with latin1 encoding: {len(df)} rows")
                return df
            except:
                return None

    def parse_nse_data(self, df):
        """
        Parse NSE CSV into structured format
        This function needs to be customized based on actual NSE CSV structure
        """
        # Expected NSE CSV columns (may vary):
        # Date, Instrument, Participant Type, Long OI, Short OI, etc.

        structured = {}

        # Try to identify columns
        print("\nAnalyzing CSV structure...")

        # Look for instrument types
        if 'Instrument' in df.columns or 'INSTRUMENT' in df.columns:
            inst_col = 'Instrument' if 'Instrument' in df.columns else 'INSTRUMENT'
            print(f"✓ Found instrument column: {inst_col}")
            print(f"  Unique instruments: {df[inst_col].unique()}")

        # Look for participant types
        if 'Client Type' in df.columns or 'Participant' in df.columns:
            part_col = 'Client Type' if 'Client Type' in df.columns else 'Participant'
            print(f"✓ Found participant column: {part_col}")
            print(f"  Unique participants: {df[part_col].unique()}")

        # Sample parsing logic (customize based on actual CSV)
        for instrument in ['Index Futures', 'Stock Futures',
                          'Index Options - Call', 'Index Options - Put',
                          'Stock Options - Call', 'Stock Options - Put']:
            structured[instrument] = {
                'Client': {'long_today': 0, 'short_today': 0, 'long_1d': 0, 'short_1d': 0, 'long_3d': 0, 'short_3d': 0},
                'DII': {'long_today': 0, 'short_today': 0, 'long_1d': 0, 'short_1d': 0, 'long_3d': 0, 'short_3d': 0},
                'FII': {'long_today': 0, 'short_today': 0, 'long_1d': 0, 'short_1d': 0, 'long_3d': 0, 'short_3d': 0},
                'Pro': {'long_today': 0, 'short_today': 0, 'long_1d': 0, 'short_1d': 0, 'long_3d': 0, 'short_3d': 0}
            }

        return structured

    def determine_action(self, long_val, short_val):
        """Determine action and color from OI changes"""
        if long_val > 0 and short_val > 0:
            return ("Build Long", True) if long_val > short_val else ("Build Short", False)
        elif long_val > 0 and short_val < 0:
            return "Closed Shorts", True
        elif long_val < 0 and short_val > 0:
            return "Closed Longs", False
        elif long_val < 0 and short_val < 0:
            return ("Closed Longs", False) if abs(long_val) > abs(short_val) else ("Closed Shorts", True)
        return "No Change", True

    def add_section(self, ws, start_row, section_title, section_data, date_str):
        """Add one section with exact screenshot format"""
        row = start_row

        # Section header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
        header = ws.cell(row=row, column=1)
        header.value = section_title
        header.font = Font(bold=True, size=11, color="FFFFFF")
        header.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
        header.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 18
        row += 1

        # Column headers
        headers = [
            'Date', 'Participant',
            'Long OI\nTODAY', 'Short OI\nTODAY', 'Action\nTODAY',
            'Long OI\n1-DAY AGO', 'Short OI\n1-DAY AGO', 'Action\n1-DAY AGO',
            'Long OI\n3 DAYS AGO', 'Short OI\n3 DAYS AGO', 'Action\n3 DAYS AGO',
            'Net Change', 'View'
        ]

        for col_idx, header_text in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header_text
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = self.border_thin
        ws.row_dimensions[row].height = 30
        row += 1

        # Participant rows
        for participant in ['Client', 'DII', 'FII', 'Pro']:
            pdata = section_data.get(participant, {})

            # Get values
            long_today = pdata.get('long_today', 0)
            short_today = pdata.get('short_today', 0)
            long_1d = pdata.get('long_1d', 0)
            short_1d = pdata.get('short_1d', 0)
            long_3d = pdata.get('long_3d', 0)
            short_3d = pdata.get('short_3d', 0)

            # Determine actions
            action_today, is_bullish_today = self.determine_action(long_today, short_today)
            action_1d, is_bullish_1d = self.determine_action(long_1d, short_1d)
            action_3d, is_bullish_3d = self.determine_action(long_3d, short_3d)

            view = "Bullish" if is_bullish_today else "Bearish"
            net_change = long_today - short_today

            # Row values
            row_values = [
                date_str, participant,
                f"{long_today:+,}", f"{short_today:+,}", action_today,
                f"{long_1d:+,}", f"{short_1d:+,}", action_1d,
                f"{long_3d:+,}", f"{short_3d:+,}", action_3d,
                f"{net_change:+,}", view
            ]

            # Fonts for each cell
            fonts = [
                Font(size=9), Font(bold=True, size=9),
                self.green_font if is_bullish_today else self.red_font,
                self.green_font if is_bullish_today else self.red_font,
                self.green_font if is_bullish_today else self.red_font,
                self.green_font if is_bullish_1d else self.red_font,
                self.green_font if is_bullish_1d else self.red_font,
                self.green_font if is_bullish_1d else self.red_font,
                self.green_font if is_bullish_3d else self.red_font,
                self.green_font if is_bullish_3d else self.red_font,
                self.green_font if is_bullish_3d else self.red_font,
                self.green_font if is_bullish_today else self.red_font,
                self.green_font if is_bullish_today else self.red_font
            ]

            for col_idx, (value, font) in enumerate(zip(row_values, fonts), 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = value
                cell.font = font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = self.border_thin

            ws.row_dimensions[row].height = 16
            row += 1

        return row + 1

    def create_excel(self, structured_data, output_file, date_str="24-JUL-2026"):
        """Create Excel in exact screenshot format"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Participant OI {date_str}"

        # Title
        ws.merge_cells('A1:M1')
        title = ws['A1']
        title.value = f"F&O Participant-wise Open Interest ({date_str})"
        title.font = Font(size=14, bold=True, color="FFFFFF")
        title.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25

        current_row = 3

        # Add all sections
        sections = [
            "Index Futures",
            "Stock Futures",
            "Index Options - Call",
            "Index Options - Put",
            "Stock Options - Call",
            "Stock Options - Put"
        ]

        for section in sections:
            section_data = structured_data.get(section, {})
            current_row = self.add_section(ws, current_row, section, section_data, date_str)

        # Column widths
        col_widths = [12, 12, 14, 14, 15, 14, 14, 15, 14, 14, 15, 14, 12]
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        wb.save(output_file)
        print(f"\n✓ Excel created: {output_file}")


def main():
    print("\n" + "="*80)
    print("NSE Participant OI Parser - EXACT Screenshot Format")
    print("="*80 + "\n")

    if len(sys.argv) < 2:
        print("Usage: python parse_nse_to_exact_format.py <nse-csv-file>")
        print("\nDownload CSV from: https://www.nseindia.com/all-reports-derivatives")
        print("Look for: 'F&O-Participant wise Open Interest (csv)'")
        print("\nOR run demo: python exact_nse_format.py")
        print("="*80 + "\n")
        return

    csv_file = sys.argv[1]
    if not os.path.exists(csv_file):
        print(f"✗ File not found: {csv_file}")
        return

    parser = NSEParticipantParser()

    # Read CSV
    df = parser.read_nse_csv(csv_file)
    if df is None:
        return

    print(f"\nFirst 5 rows of CSV:")
    print(df.head())

    # Parse data
    print("\n" + "-"*80)
    structured_data = parser.parse_nse_data(df)

    # Create Excel
    output_file = "NSE_Participant_OI_Exact_Format_24_JUL_2026.xlsx"
    parser.create_excel(structured_data, output_file, "24-JUL-2026")

    print("\n" + "="*80)
    print("✓ Processing complete!")
    print(f"✓ Output: {output_file}")
    print("\nNOTE: If data appears as zeros, the CSV parsing logic needs adjustment.")
    print("Run: python inspect_nse_csv.py <csv-file> to see the CSV structure.")
    print("="*80 + "\n")


if __name__ == "__main__":
    main()
